import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pygame
import time
import re
import sys
import os


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


os.environ["SDL_AUDIODRIVER"] = "directsound"
try:
    pygame.mixer.init()
    pygame.mixer.music.load(resource_path("barka.mp3"))
    pygame.mixer.music.play(-1)
except pygame.error as e:
    print("Błąd audio, pomijam dźwięk:", e)

# =========================
# ===== SELENIUM ==========
# =========================
def init_driver():
    options = uc.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-service-autorun")
    options.add_argument("--password-store=basic")
    # options.add_argument("--headless")   # <-- odkomentuj, jeśli nie chcesz widzieć okna
    driver = uc.Chrome(options=options, version_main=0)
    driver.set_page_load_timeout(10)

    wait = WebDriverWait(driver, 3)

    driver.get("https://www.dell.com/support/home/pl-pl")

    try:
        cookie_btn = wait.until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )
        cookie_btn.click()
    except Exception:
        pass
    return driver


def get_dell_warranty(driver, service_tag):
    wait = WebDriverWait(driver, 3)

    search = wait.until(
        EC.presence_of_element_located((By.ID, "mh-search-input"))
    )
    try:
        old_warranty_elem = driver.find_element(By.ID, "tt_warstatus_text")
    except Exception:
        old_warranty_elem = None

    search.clear()
    search.send_keys(service_tag)
    search.send_keys(Keys.ENTER)

    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "h1")))

    if old_warranty_elem is not None:
        wait.until(EC.staleness_of(old_warranty_elem))

    data = {"Serial": service_tag, "Gwarancja": None, "ZlyST": False}
    try:
        warr_div = wait.until(
            EC.presence_of_element_located((By.ID, "tt_warstatus_text"))
        )
        value = warr_div.get_attribute("innerText").strip()
        if value:
            data["Gwarancja"] = value
        else:
            data["ZlyST"] = True
    except Exception:
        data["ZlyST"] = True

    return data


def run_selenium(service_tags):
    results = []
    driver = init_driver()
    try:
        for tag in service_tags:
            print(f"➡ Sprawdzam gwarancję dla: {tag}")
            try:
                result = get_dell_warranty(driver, tag)
            except Exception as e:
                print(f"❌ Błąd przy {tag}: {e}")
                result = {"Serial": tag, "Gwarancja": None, "ZlyST": True}
            results.append(result)
            time.sleep(1.5)            # krótkie opóźnienie
    finally:
        driver.quit()
    return results


def save_excel(data, file="dell_warranty.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ST", "Gwarancja"])
    for row in data:
        if row.get("ZlyST"):
            ws.append([row["Serial"], "BŁĘDNY SERVICE TAG"])
        else:
            ws.append([row["Serial"], row["Gwarancja"]])
    wb.save(file)
    print(f"📁 Dane zapisane do: {file}")

# =========================
# ===== GUI ===============
# =========================
service_tags = []
root = tk.Tk()
root.title("Kamil Dell wariat Checker")
root.geometry("420x450")
bg_image = tk.PhotoImage(file=resource_path("jp2.gif"))
bg_label = tk.Label(root, image=bg_image)
bg_label.place(x=0, y=60, relwidth=1, relheight=1)

tk.Label(root, text="Wpisz wariaty recznie, spacja, przecinek, srednik pomiedzy:").pack(pady=5)
entry_st = tk.Entry(root, width=40)
entry_st.pack(pady=5)

listbox = tk.Listbox(root, width=40, height=10)
listbox.pack(pady=10)

label_count = tk.Label(root, text="Łącznie wariatow: 0")
label_count.pack(pady=5)


def update_count():
    label_count.config(text=f"Łącznie wariatow: {len(service_tags)}")


def add_st():
    raw = entry_st.get().strip()
    if not raw:
        return
    tags = [st.strip() for st in re.split(r"[,\s;]+", raw) if st.strip()]
    for st in tags:
        if st not in service_tags:
            service_tags.append(st)
            listbox.insert(tk.END, st)
    entry_st.delete(0, tk.END)
    update_count()


def load_from_excel():
    path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if not path:
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if "ST" not in headers:
        messagebox.showerror("Błąd", "Brak kolumny 'ST'")
        return
    col_index = headers.index("ST")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_index]:
            st = str(row[col_index]).strip()
            if st and st not in service_tags:
                service_tags.append(st)
                listbox.insert(tk.END, st)
    update_count()


def start_check():
    if not service_tags:
        messagebox.showwarning("Brak danych", "Dodaj Service Tagi")
        return
    results = run_selenium(service_tags)
    save_excel(results)
    messagebox.showinfo("Gotowe", "Zapisano dell_warranty.xlsx")


tk.Button(root, text="Dodaj wariaty recznie", command=add_st).pack(pady=5)
tk.Button(root, text="Wczytaj wariaty z XLSX", command=load_from_excel).pack(pady=5)
tk.Label(root, text="Zczyta wariaty z .xlsx z kolumny ST").pack(pady=5)
tk.Button(root, text="START", command=start_check, bg="green", fg="white").pack(pady=10)

root.mainloop()